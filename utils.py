from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from requests import get
from os import path, makedirs
from re import findall


class Date:
    """日期类"""

    def __init__(self) -> None:
        self.month = datetime.now().month  # 月份
        self.day = datetime.now().day  # 天数

    @property
    def get_date(self):
        """
        获取日期（格式：x月x日）
        """

        return '{}月{}日'.format(self.month, self.day)


def get_coord(point: str) -> (int, int):
    """
    根据提供的字符串获取单元格坐标
    :param point: 形如 "C12"
    :return: 形如 (12,3)
    """
    alpha_list = []
    digit_list = []

    for ch in point:  # 分离字符串
        if ch.isalpha():
            alpha_list.append(ch)
        elif ch.isdigit():
            digit_list.append(ch)
        else:  # 出现其他字符
            return TypeError

    row = ''.join(digit_list)
    row = int(row)

    col = ''.join(alpha_list)
    col = column_index_from_string(col)

    return row, col


class ExcelFile:
    """Excel文件类"""

    def __init__(self, file_path: str) -> None:
        """
        初始化
        :param file_path: 文件路径
        """
        index = file_path.rindex(".")
        self.file_path = file_path  # Excel文件路径
        self.suffix = file_path[index + 1:]  # Excel文件后缀


class Excel2Img(ExcelFile):
    """Excel链接提取类"""

    def __init__(self, file_path: str, dir_path: str, sheet_id: int = 1):
        """
        传参
        :param dir_path: 导入文件夹路径
        :param file_path: 导出文件路径
        :param sheet_id: 工作表 id
        """
        super().__init__(file_path)
        self.dir_path = dir_path  # 导入文件夹路径
        self.book = load_workbook(self.file_path)  # 读取表格
        self.sheet = self.book.worksheets[sheet_id - 1]  # 读取工作表

    def set_sheet(self, sheet_id):
        self.sheet = self.book.worksheets[sheet_id - 1]  # 读取工作表

    def get_preview(self, name_rule: str) -> str:
        cell_name_list = findall(r"\$(.*?)\$", name_rule)

        temp_name = name_rule
        for cell_name in cell_name_list:
            old_str = '$' + cell_name + '$'
            try:
                point = get_coord(cell_name)
                temp_name = temp_name.replace(old_str, self.sheet.cell(row=point[0], column=point[1]).value)
            except:
                temp_name = temp_name.replace(old_str, cell_name)

        return temp_name

    def get_names(self, name_rule: str, count: int) -> list[str]:
        """
        获取Excel文件中，根据命名规则进行命名的字符串规则
        :param name_rule: 命名规则
        :param count: 名称数量
        """
        cell_name_list = findall(r"\$(.*?)\$", name_rule)

        names = []

        for i in range(count):
            temp_name = name_rule
            for cell_name in cell_name_list:
                old_str = '$' + cell_name + '$'
                try:
                    point = get_coord(cell_name)
                    temp_name = temp_name.replace(old_str, self.sheet.cell(row=point[0] + i, column=point[1]).value)
                except:
                    temp_name = temp_name.replace(old_str, cell_name)
            names.append(temp_name)

        return names

    def get_urls(self, start=(1, 1), count: int = 0) -> list[str]:
        """
        获取Excel文件中，从start单元格起始向下count个单元格内的超链接
        :param start: 起始单元格 (r,c)
        :param count: 数量
        :return: 链接（str）列表
        """
        urls = []
        # 读取超链接
        col = start[1]
        for row in range(start[0], start[0] + count):
            urls.append(self.sheet.cell(row, col).hyperlink.target)
        return urls

        # # 识别后缀，分别调用xlrd(xls)，xlsx(openpyxl)相应算法
        # if self.suffix == "xls":
        #     # 获取excel表
        #     main_book = open_workbook(self.path, formatting_info=True)
        #     main_sheet = main_book.sheet_by_index(0)
        #
        #     # 读取超链接
        #     urls = []
        #     for i in range(start[0], start[0] + count):
        #         for j in range(start[1], start[1] + count):
        #             urls.append(main_sheet.hyperlink_map.get((i - 1, j - 1)).url_or_path)
        #     return urls
        # else:
        #     # 获取excel表
        #     main_book = load_workbook(self.path)
        #     main_sheet = main_book.active
        #
        #     # 读取超链接
        #     urls = []
        #     for i in range(start[0], start[0] + count):
        #         for j in range(start[1], start[1] + count):
        #             urls.append(main_sheet.cell(i, j).hyperlink.target)
        #     return urls

    def excel2img(self, name_rule: str, point: str, count: int = 0, is_date: bool = True):
        date = Date().get_date

        save_path = self.dir_path

        start = get_coord(point)  # 获取坐标

        # 获取链接
        urls = self.get_urls(start, count)
        names = self.get_names(name_rule, count)
        if is_date:
            save_path += '/{}'.format(date)

        tmp = save_path + "/" + names[0]
        index = tmp.rindex("/")
        tmp = tmp[:index]
        if not path.exists(tmp):
            makedirs(tmp)

        # 把下载地址发送给requests模块
        for i in range(count):
            name = names[i]

            file = get(urls[i])  # 下载文件
            index_file = urls[i].rindex('=')
            suffix_file = urls[i][index_file + 1:]
            # 下载文件
            with open("{}/{}.{}".format(save_path, name, suffix_file), "wb") as code:  # 写入
                code.write(file.content)

        return '转换成功'

# if __name__ == '__main__':
#     ex = ExcelFile("./excel_test.xlsx")
#     # direct = input('请输入保存的文件夹名称：')
#     # tool = Excel2Img(direct)
#     # print(tool.excel2img('excel.xlsm'))
